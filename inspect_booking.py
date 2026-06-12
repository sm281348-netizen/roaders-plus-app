
import openpyxl, os, sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

filepath = r"Y:\櫃台\金旭\A1自動入房租前核對表\2026\202606\20260604.XLSX"
wb = openpyxl.load_workbook(filepath, data_only=True)
print(f"工作表: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ({ws.max_row}列 x {ws.max_column}欄) ===")
    for r in range(1, min(ws.max_row + 1, 30)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is not None and val != "":
                row_vals.append(f"[{c}]{repr(val)}")
        if row_vals:
            print(f"  Row{r:02d}: {' | '.join(row_vals)}")
